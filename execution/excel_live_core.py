import os
from dataclasses import dataclass
from typing import Dict, Any

from openpyxl import load_workbook


# -------------------------------------------------
# INPUT STRUCTURE
# -------------------------------------------------

@dataclass
class CoreInputs:

    trend_strength: float
    structure_ok: bool
    volume_score: float
    risk_state: str
    confidence_score: float
    volatility_regime: str


# -------------------------------------------------
# EXCEL AI CORE
# -------------------------------------------------

class ExcelLiveCore:

    def __init__(self, excel_path: str):

        if not os.path.exists(excel_path):
            raise FileNotFoundError(excel_path)

        self.excel_path = excel_path

        # Excel sheets
        self.input_sheet = "PYTHON_BRIDGE"
        self.output_sheet = "AI_OUTPUT"

        # AI execution threshold
        self.execute_threshold = 0.6


    # -------------------------------------------------
    # VOLATILITY REGIME MAPPING
    # -------------------------------------------------

    def _volatility_to_numeric(self, regime: str) -> int:

        mapping = {
            "LOW": 0,
            "MEDIUM": 1,
            "HIGH": 2
        }

        return mapping.get(regime.upper(), 1)


    # -------------------------------------------------
    # WRITE INPUTS TO EXCEL
    # -------------------------------------------------

    def _write_inputs(self, inputs: CoreInputs):

        wb = load_workbook(self.excel_path)

        ws = wb[self.input_sheet]

        values = {
            "confidence_score_input": inputs.confidence_score,
            "volume_score_input": inputs.volume_score,
            "trend_strength_input": inputs.trend_strength,
            "volatility_regime_input": self._volatility_to_numeric(inputs.volatility_regime),
        }

        for row in ws.iter_rows(min_row=2):

            field = row[0].value

            if field in values:
                row[1].value = values[field]

        # ---------------------------------------------
        # FORCE EXCEL RECALCULATION
        # ---------------------------------------------

        wb.calculation.fullCalcOnLoad = True

        wb.save(self.excel_path)


    # -------------------------------------------------
    # READ AI SCORE
    # -------------------------------------------------

    def _read_score(self) -> float:

        wb = load_workbook(self.excel_path, data_only=True)

        ws = wb[self.output_sheet]

        score = ws["B2"].value

        if score is None:
            return 0.0

        try:
            return round(float(score), 4)
        except:
            return 0.0


    # -------------------------------------------------
    # FINAL DECISION
    # -------------------------------------------------

    def decide(self, inputs: CoreInputs) -> Dict[str, Any]:

        # write inputs to excel
        self._write_inputs(inputs)

        # read calculated score
        ai_score = self._read_score()

        decision = "EXECUTE" if ai_score >= self.execute_threshold else "BLOCK"

        return {
            "ai_score": ai_score,
            "final_trade_decision": decision
        }
